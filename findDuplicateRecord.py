# 小伙伴希望能依据姓名和手机号来对两张表查重，并找出互相的差集

from openpyxl import load_workbook
from itertools import islice
import pandas as pd

# wbA 转为dfA
wbA = load_workbook("1.xlsx")
sheetA = wbA.get_sheet_by_name("Sheet1")
dataA = sheetA.values
colsA = next(dataA)[0:]
dataA = list(dataA)
# dataA = (islice(r, 1, None) for r in dataA)
dfA = pd.DataFrame(dataA, index=None, columns=colsA)

# wbB 转为dfB
wbB = load_workbook("2.xlsx")
sheetB = wbB.get_sheet_by_name("Sheet1")
dataB = sheetB.values
colsB = next(dataB)[0:]
dataB = list(dataB)
# dataB = (islice(r, 1, None) for r in dataB)
dfB = pd.DataFrame(dataB, index=None, columns=colsB)
# print(dfB)

# 表中"报名人真实姓名"，"手机号"作为一列
dfA["报名人真实姓名"]=dfA["报名人真实姓名"].map(str.strip)
dfA["手机号"]=dfA["手机号"].map(str.strip)
dfB["报名人真实姓名"]=dfB["报名人真实姓名"].map(str.strip)
dfB["手机号"]=dfB["手机号"].map(str.strip)
dfA["identify"] = dfA["报名人真实姓名"]+dfA["手机号"]
dfB["identify"] = dfB["报名人真实姓名"]+dfB["手机号"]

# 查重，去重
IsDuplicated = dfB["identify"].duplicated()
dfB["IsDuplicated"] = IsDuplicated
print(dfB[dfB["IsDuplicated"]==True])

dfA = dfA.drop_duplicates(["identify"])
dfB = dfB.drop_duplicates(["identify"])

# 找到A，B集合互相的差
setA = set(dfA["identify"])
setB = set(dfB["identify"])
print("只属于集合A，不属于集合B：")
print(setA - setB)
print("只属于集合B，不属于集合A：")
print(len(setB - setA))
